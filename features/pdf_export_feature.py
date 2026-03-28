import os

import fitz  # PyMuPDF
import markdown
from pdf2docx import Converter
from PySide6.QtWidgets import (
    QVBoxLayout, QGridLayout, QLabel,
    QScrollArea, QWidget
)

from .base_feature import BaseFeature, AppCard

_COLOR_WORD  = "#2B579A"
_COLOR_EXCEL = "#217346"
_COLOR_IMG   = "#D97706"
_COLOR_MD_IN = "#B5591A"


# ── Conversion functions ───────────────────────────────────────────────────────

def _do_word(src: str, dest: str):
    cv = Converter(src)
    cv.convert(dest, start=0, multi_processing=False)
    cv.close()


def _do_excel(src: str, dest: str):
    import pdfplumber
    import openpyxl

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    with pdfplumber.open(src) as pdf:
        if not pdf.pages:
            raise ValueError("The PDF has no pages.")
        for i, page in enumerate(pdf.pages):
            ws = wb.create_sheet(title=f"Page {i + 1}")
            tables = page.extract_tables()
            if tables:
                row_num = 1
                for table in tables:
                    for row in table:
                        for col, cell in enumerate(row, 1):
                            ws.cell(row=row_num, column=col, value=cell or "")
                        row_num += 1
                    row_num += 1
            else:
                for r, line in enumerate((page.extract_text() or "").splitlines(), 1):
                    ws.cell(row=r, column=1, value=line)

    wb.save(dest)


def _do_images_to_pdf(images: list, dest: str):
    doc = fitz.open()
    for img_path in images:
        img_doc = fitz.open(img_path)
        pdf_bytes = img_doc.convert_to_pdf()
        img_doc.close()
        img_pdf = fitz.open("pdf", pdf_bytes)
        doc.insert_pdf(img_pdf)
        img_pdf.close()
    doc.save(dest)
    doc.close()


def _do_md_to_pdf(src: str, dest: str):
    with open(src, "r", encoding="utf-8") as f:
        md_text = f.read()
    html_body = markdown.markdown(md_text, extensions=["tables", "fenced_code"])
    html = f"""<!DOCTYPE html>
<html><head><meta charset="utf-8">
<style>
  body {{ font-family: Arial, sans-serif; font-size: 12pt; line-height: 1.6; margin: 40px; }}
  h1, h2, h3 {{ color: #222; }}
  code {{ background: #f4f4f4; padding: 2px 4px; border-radius: 3px; font-family: monospace; }}
  pre {{ background: #f4f4f4; padding: 10px; border-radius: 5px; }}
  table {{ border-collapse: collapse; width: 100%; }}
  th, td {{ border: 1px solid #ccc; padding: 6px 10px; }}
  th {{ background: #eee; }}
  hr {{ border: none; border-top: 1px solid #ccc; }}
</style>
</head><body>{html_body}</body></html>"""
    story = fitz.Story(html=html)
    writer = fitz.DocumentWriter(dest)
    mediabox = fitz.paper_rect("a4")
    margin = 50
    where = mediabox + (margin, margin, -margin, -margin)
    more = 1
    while more:
        dev = writer.begin_page(mediabox)
        more, _ = story.place(where)
        story.draw(dev)
        writer.end_page()
    writer.close()


# ── Feature widget ─────────────────────────────────────────────────────────────

class PdfExportFeature(BaseFeature):
    NAV_NAME = "Converter"

    def __init__(self):
        super().__init__()
        self._setup_ui()

    def _setup_ui(self):
        scroll = QScrollArea(self)
        scroll.setWidgetResizable(True)
        scroll.setStyleSheet("QScrollArea { border:none; background:transparent; }")

        container = QWidget()
        container.setStyleSheet("background:transparent;")
        scroll.setWidget(container)

        outer = QVBoxLayout(self)
        outer.setContentsMargins(0, 0, 0, 0)
        outer.addWidget(scroll)

        layout = QVBoxLayout(container)
        layout.setContentsMargins(40, 40, 40, 40)
        layout.setSpacing(20)

        layout.addWidget(QLabel("<h1>Converter</h1>"))
        layout.addWidget(QLabel(
            "Select a format — you'll be prompted to choose your file."
        ))

        grid = QGridLayout()
        grid.setSpacing(16)
        grid.setColumnStretch(0, 1)
        grid.setColumnStretch(1, 1)
        grid.setColumnStretch(2, 1)

        cards = [
            AppCard("W",   _COLOR_WORD,  "PDF to Word",
                "Convert to an editable .docx document.",
                self._convert_word),
            AppCard("X",   _COLOR_EXCEL, "PDF to Excel",
                "Tables extracted. Falls back to plain text.",
                self._convert_excel),
            AppCard("IMG", _COLOR_IMG,   "Images to PDF",
                "Combine JPG, PNG or other images into a PDF.",
                self._convert_images_to_pdf),
            AppCard("MD",  _COLOR_MD_IN, "Markdown to PDF",
                "Styled PDF with tables and code blocks.",
                self._convert_md_to_pdf),
        ]

        for i, card in enumerate(cards):
            grid.addWidget(card, i // 3, i % 3)

        layout.addLayout(grid)
        layout.addStretch()

    # ── Handlers ──────────────────────────────────────────────────────────────

    def _stem(self, path: str) -> str:
        return os.path.splitext(path)[0]

    def _convert_word(self):
        src = self.open_file_dialog("Select a PDF", "PDF (*.pdf)")
        if not src:
            return
        dest = self.save_dialog("Save Word Document", self._stem(src) + ".docx", "Word (*.docx)")
        if dest:
            self.run_with_dialog(lambda: _do_word(src, dest),
                                 f"Converted!\n{os.path.basename(dest)}")

    def _convert_excel(self):
        src = self.open_file_dialog("Select a PDF", "PDF (*.pdf)")
        if not src:
            return
        dest = self.save_dialog("Save Excel", self._stem(src) + ".xlsx", "Excel (*.xlsx)")
        if dest:
            self.run_with_dialog(lambda: _do_excel(src, dest),
                                 f"Converted!\n{os.path.basename(dest)}")

    def _convert_images_to_pdf(self):
        images = self.open_files_dialog(
            "Select Images",
            "Images (*.jpg *.jpeg *.png *.bmp *.webp *.tiff *.tif)"
        )
        if not images:
            return
        dest = self.save_dialog("Save PDF", "images.pdf", "PDF (*.pdf)")
        if not dest:
            return
        self.run_with_dialog(
            lambda: _do_images_to_pdf(images, dest),
            f"Converted {len(images)} image{'s' if len(images) != 1 else ''} to:\n{os.path.basename(dest)}"
        )

    def _convert_md_to_pdf(self):
        src = self.open_file_dialog("Select a Markdown file", "Markdown (*.md *.markdown)")
        if not src:
            return
        dest = self.save_dialog("Save PDF", self._stem(src) + ".pdf", "PDF (*.pdf)")
        if dest:
            self.run_with_dialog(lambda: _do_md_to_pdf(src, dest),
                                 f"Converted!\n{os.path.basename(dest)}")
